"""Wspólne helpery stylowania eksportów Excela (openpyxl), zero streamlit.

Konsoliduje kod powielony wcześniej w 4 miejscach (convert_df, eksport_karty_maski,
generuj_excel_rejestru, eksport w check_sales.py). Nie każdy eksport używa każdego
helpera - część eksportów ma świadomie minimalną stylizację (patrz docstring
poszczególnych funkcji, gdzie są wołane).
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from config.constants import KOLOR_MARKI_CZERWONY, KOLOR_OBRAMOWANIA, KOLOR_ZEBRA_JASNY


def dopasuj_szerokosci_kolumn(worksheet, margines=3, min_szerokosc=12):
    """
    Autodopasowanie szerokości kolumn na podstawie długości najdłuższej wartości
    (nagłówek lub dane) w każdej kolumnie.

    Wywoływane z: convert_df (margines=3, min=12), generuj_excel_rejestru
    (margines=4, min=15), eksportu w check_sales.py (margines=3, min=14).
    Nie używa go eksport_karty_maski - tam szerokości są stałe (zależą od stałej
    liczby poziomów siatki rozmiarowej, nie od treści).
    """
    for col in worksheet.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        worksheet.column_dimensions[col_letter].width = max(max_len + margines, min_szerokosc)


def stylizuj_naglowek_marki(worksheet, wysokosc_wiersza=80):
    """
    Brandowany nagłówek eksportu: czerwone tło marki, biała pogrubiona czcionka,
    wyśrodkowanie z zawijaniem tekstu, cienkie szare obramowanie.

    Wywoływane z: convert_df (wysokosc_wiersza=80) i eksport_karty_maski
    (wysokosc_wiersza=40) w app.py. Nie używane w check_sales.py (tam nagłówek
    zostaje domyślny, bez stylizacji) ani w generuj_excel_rejestru (tam nagłówek
    ma tylko pogrubienie, bez tła i obramowania) - to świadomie różne style,
    nie jeden uniwersalny.
    """
    font_naglowek = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    fill_naglowek = PatternFill(start_color=KOLOR_MARKI_CZERWONY, end_color=KOLOR_MARKI_CZERWONY, fill_type="solid")
    align_naglowek_srodek = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_cienka = Border(
        left=Side(style='thin', color=KOLOR_OBRAMOWANIA),
        right=Side(style='thin', color=KOLOR_OBRAMOWANIA),
        top=Side(style='thin', color=KOLOR_OBRAMOWANIA),
        bottom=Side(style='thin', color=KOLOR_OBRAMOWANIA),
    )

    worksheet.row_dimensions[1].height = wysokosc_wiersza
    for cell in worksheet[1]:
        cell.font = font_naglowek
        cell.fill = fill_naglowek
        cell.alignment = align_naglowek_srodek
        cell.border = border_cienka


def stylizuj_wiersze_danych(worksheet, kolumny_wyrownane_w_lewo, kolumna_pogrubiona=1):
    """
    Styl wierszy danych (od wiersza 2 w dół) w brandowanym eksporcie: cienkie
    obramowanie, naprzemienne jasne tło (zebra) na parzystych wierszach, wysokość
    wiersza 20, wyrównanie do lewej dla podanych indeksów kolumn (reszta
    wyśrodkowana), pogrubienie wskazanej kolumny (np. Indeksu).

    Wywoływane z: convert_df (kolumny_wyrownane_w_lewo wyliczone dynamicznie:
    1, 2, 3 oraz kolumna "Skąd zabrać"; kolumna_pogrubiona=1) i eksport_karty_maski
    (kolumny_wyrownane_w_lewo=[2, 3]; kolumna_pogrubiona=2).

    Nie obejmuje alarmowego podświetlania dużego złogu w convert_df - to reguła
    biznesowa tego jednego raportu, dokładana przez convert_df jako osobny krok
    PO wywołaniu tego helpera (nadpisuje fill/font na konkretnych komórkach).
    """
    font_dane = Font(name="Segoe UI", size=10, bold=False)
    font_pogrubiona = Font(name="Segoe UI", size=10, bold=True)
    fill_paski = PatternFill(start_color=KOLOR_ZEBRA_JASNY, end_color=KOLOR_ZEBRA_JASNY, fill_type="solid")
    border_cienka = Border(
        left=Side(style='thin', color=KOLOR_OBRAMOWANIA),
        right=Side(style='thin', color=KOLOR_OBRAMOWANIA),
        top=Side(style='thin', color=KOLOR_OBRAMOWANIA),
        bottom=Side(style='thin', color=KOLOR_OBRAMOWANIA),
    )
    align_lewa = Alignment(horizontal="left", vertical="center", wrap_text=False)
    align_srodek = Alignment(horizontal="center", vertical="center", wrap_text=False)

    max_row = worksheet.max_row
    max_col = worksheet.max_column

    for row_idx in range(2, max_row + 1):
        worksheet.row_dimensions[row_idx].height = 20
        for col_idx in range(1, max_col + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.font = font_pogrubiona if col_idx == kolumna_pogrubiona else font_dane
            cell.border = border_cienka
            cell.alignment = align_lewa if col_idx in kolumny_wyrownane_w_lewo else align_srodek
            if row_idx % 2 == 0:
                cell.fill = fill_paski


def dodaj_tabele_stylu_lekkiego(worksheet, nazwa_tabeli):
    """
    Owija cały zakres danych arkusza w formalny obiekt Table Excela ze stylem
    "TableStyleLight1" (te same paski/obramowanie natywnej tabeli Excela w każdym
    brandowanym eksporcie). Zakres liczony automatycznie z max_row/max_column
    arkusza (zakłada, że dane zaczynają się w A1).

    Wywoływane z: convert_df ("TabelaRekomendacji") i eksport_karty_maski
    ("TabelaSiatkiRozmiarowej") w app.py.
    """
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    ostatnia_litera = get_column_letter(max_col)

    tabela = Table(displayName=nazwa_tabeli, ref=f"A1:{ostatnia_litera}{max_row}")
    tabela.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    worksheet.add_table(tabela)
