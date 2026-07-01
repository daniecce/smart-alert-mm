"""Verifier screen: upload a past transfer history file and check its sales effectiveness."""

import streamlit as str_web
import pandas as pd
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl.styles import Font, Alignment, Border, Side

from core.excel_export import dopasuj_szerokosci_kolumn

def wyswietl_weryfikator_sprzedazy():
    # --- PRZYCISK POWROTU ---
    if str_web.button("◀ POWRÓT DO GŁÓWNEGO RAPORTU"):
        str_web.session_state['aktywny_ekran'] = "glowny"
        str_web.rerun()

    str_web.subheader("📊 Weryfikator Przesunięć")
    str_web.markdown(
        "Wgraj plik z historią przesunięć (kolumny: **Indeks**, **Skąd**, **Dokąd**, opcjonalnie: *Dni na filii*), "
        "aby sprawdzić skuteczność alokacji, finanse oraz czas zalegania przed wykonaniem MM."
    )

    # --- SEKCJA PLIKU I DAT ---
    col1, col2 = str_web.columns([1, 1])

    with col1:
        plik_wgrany = str_web.file_uploader("Wgraj plik Excel z przesunięciami", type=['xlsx', 'xls', 'csv'])

    with col2:
        dzisiaj = datetime.now().date()
        domyslny_start = dzisiaj - timedelta(days=10)

        zakres_dat = str_web.date_input(
            "Wybierz zakres dat do analizy sprzedaży w sklepie docelowym:",
            value=(domyslny_start, dzisiaj),
            max_value=dzisiaj
        )

    if plik_wgrany is not None:
        if len(zakres_dat) != 2:
            str_web.warning("Proszę wybrać pełny zakres dat (od - do) na kalendarzu.")
            str_web.stop()

        data_start, data_koniec = zakres_dat

        try:
            if plik_wgrany.name.endswith('.csv'):
                df_wgrany = pd.read_csv(plik_wgrany)
            else:
                df_wgrany = pd.read_excel(plik_wgrany)
        except Exception as e:
            str_web.error(f"Błąd podczas odczytu pliku: {e}")
            str_web.stop()

        # Weryfikacja podstawowej struktury pliku
        wymagane_kolumny = ['Indeks', 'Skąd', 'Dokąd']
        brakujace = [kol for kol in wymagane_kolumny if kol not in df_wgrany.columns]

        if brakujace:
            str_web.error(
                f"W pliku brakuje wymaganych kolumn: {', '.join(brakujace)}. "
                f"Upewnij się, że nazwy to dokładnie: Indeks, Skąd, Dokąd."
            )
            str_web.stop()

        # --- CZYSZCZENIE I AGREGACJA PLIKU WEJŚCIOWEGO ---
        df_wgrany['Indeks'] = df_wgrany['Indeks'].astype(str).str.strip()
        df_wgrany['Skąd'] = df_wgrany['Skąd'].astype(str).str.strip()
        df_wgrany['Dokąd'] = df_wgrany['Dokąd'].astype(str).str.strip()

        # Ujednolicona nazwa kolumny
        nazwa_kolumny_zlogu = 'Dni na filii'

        # Dynamicznie sprawdzamy, czy kolumna istnieje w pliku pod nową nazwą
        if nazwa_kolumny_zlogu in df_wgrany.columns:
            df_wgrany[nazwa_kolumny_zlogu] = pd.to_numeric(df_wgrany[nazwa_kolumny_zlogu], errors='coerce').fillna(0).astype(int)
            ma_kolumne_zlogu = True
        else:
            ma_kolumne_zlogu = False

        # Grupowanie przy użyciu Named Aggregation dla pełnego bezpieczeństwa nazw
        df_wgrany_skonsolidowany = df_wgrany.groupby(['Indeks', 'Dokąd']).agg(**{
            'Skąd': pd.NamedAgg(column='Skąd', aggfunc=lambda x: ", ".join(sorted(list(set(x))))),
            'Ile (MM)': pd.NamedAgg(column='Skąd', aggfunc='count'),
            'Dni_Zlogu': pd.NamedAgg(column=nazwa_kolumny_zlogu if ma_kolumne_zlogu else 'Indeks',
                                     aggfunc='max' if ma_kolumne_zlogu else lambda x: 0)
        }).reset_index()

        df_wgrany_skonsolidowany['Dni_Zlogu'] = df_wgrany_skonsolidowany['Dni_Zlogu'].astype(int)

        # --- ANALIZA SPRZEDAŻY I FINANSÓW ---
        df_sprzedaz_surowa = str_web.session_state.get('df_sprzedaz_surowa')
        if df_sprzedaz_surowa is None or df_sprzedaz_surowa.empty:
            str_web.error("Brak danych sprzedażowych w pamięci systemu.")
            str_web.stop()

        with str_web.spinner("Trwa przeliczanie skuteczności i finansów..."):
            maska_czasowa = (
                (df_sprzedaz_surowa['Data_Transakcji'].dt.date >= data_start) &
                (df_sprzedaz_surowa['Data_Transakcji'].dt.date <= data_koniec)
            )
            sprzedaz_w_okresie = df_sprzedaz_surowa[maska_czasowa]

            # Agregacja danych z bazy sprzedaży
            sprzedaz_zagregowana = sprzedaz_w_okresie.groupby(['Indeks', 'Filia']).agg({
                'Ilosc_szt': 'sum',
                'Wartosc_Netto': 'sum',
                'Marza_Faktyczna': 'sum'
            }).reset_index()

            sprzedaz_zagregowana.columns = [
                'Indeks', 'Dokąd', 'Faktyczna_Sprzedaz_Szt', 'Faktyczny_Obrot_Netto', 'Faktyczna_Marza_Netto'
            ]
            sprzedaz_zagregowana['Indeks'] = sprzedaz_zagregowana['Indeks'].astype(str).str.strip()
            sprzedaz_zagregowana['Dokąd'] = sprzedaz_zagregowana['Dokąd'].astype(str).str.strip()

            # Łączymy wyniki
            wynik = pd.merge(df_wgrany_skonsolidowany, sprzedaz_zagregowana, on=['Indeks', 'Dokąd'], how='left')

            wynik['Faktyczna_Sprzedaz_Szt'] = wynik['Faktyczna_Sprzedaz_Szt'].fillna(0).astype(int)
            wynik['Faktyczny_Obrot_Netto'] = wynik['Faktyczny_Obrot_Netto'].fillna(0.0)
            wynik['Faktyczna_Marza_Netto'] = wynik['Faktyczna_Marza_Netto'].fillna(0.0)

            # Układ kolumn: 'Dni_Zlogu' od razu po kolumnie 'Indeks'
            wynik = wynik[[
                'Indeks', 'Dni_Zlogu', 'Skąd', 'Dokąd', 'Ile (MM)',
                'Faktyczna_Sprzedaz_Szt', 'Faktyczny_Obrot_Netto', 'Faktyczna_Marza_Netto'
            ]]

            # Kalkulacja podsumowań (TOTAL)
            suma_paczek_mm = wynik['Ile (MM)'].sum()
            srednie_dni_zlogu = wynik['Dni_Zlogu'].mean() if len(wynik) > 0 else 0
            suma_sprzedanych_sztuk = wynik['Faktyczna_Sprzedaz_Szt'].sum()
            total_obrot = wynik['Faktyczny_Obrot_Netto'].sum()
            total_marza = wynik['Faktyczna_Marza_Netto'].sum()

            # Liczba przesunięć z pliku źródłowego, które zakończyły się sukcesem (sprzedażą)
            przesuniecia_trafione = wynik.loc[wynik['Faktyczna_Sprzedaz_Szt'] > 0, 'Ile (MM)'].sum()

            # Skuteczność liczona od całkowitej liczby paczek MM
            skutecznosc = (przesuniecia_trafione / suma_paczek_mm * 100) if suma_paczek_mm > 0 else 0

        # --- WYŚWIETLANIE WYNIKÓW ---
        str_web.markdown("---")
        col_stat1, col_stat2, col_stat3, col_stat4, col_stat5 = str_web.columns(5)

        col_stat1.metric(
            "Wszystkie przesunięcia", f"{suma_paczek_mm} szt",
            help="Łączna liczba wszystkich pozycji/wierszy z przesunięciami przesłanych w pliku wejściowym."
        )
        col_stat2.metric(
            "Skuteczność alokacji", f"{skutecznosc:.1f}%",
            help="Procent wszystkich przesunięć (wierszy z pliku), które trafiły do sklepów i wygenerowały sprzedaż min. 1 szt."
        )
        col_stat3.metric(
            "TOTAL SPRZEDANE SZTUKI", f"{suma_sprzedanych_sztuk} szt",
            help="Suma wszystkich sztuk towaru sprzedanych we wszystkich sklepach docelowych w wybranym oknie czasowym."
        )
        col_stat4.metric(
            "TOTAL OBRÓT NETTO", f"{total_obrot:,.2f} PLN",
            help="Suma rzeczywistego obrotu netto wygenerowanego na sklepach docelowych (bez powtórzeń)."
        )
        col_stat5.metric(
            "TOTAL MARŻA NETTO", f"{total_marza:,.2f} PLN",
            help="Suma rzeczywistego zysku na marży zarobionego na czysto przez te modele."
        )

        str_web.markdown("<br>", unsafe_allow_html=True)

        str_web.dataframe(
            wynik,
            width="stretch",
            hide_index=True,
            column_config={
                "Dni_Zlogu": str_web.column_config.NumberColumn("Dni na filii", format="%d dni" if ma_kolumne_zlogu else "-", alignment="center"),
                "Ile (MM)": str_web.column_config.NumberColumn("Ile (MM)", format="%d szt", alignment="center"),
                "Faktyczna_Sprzedaz_Szt": str_web.column_config.NumberColumn("Sprzedaż (szt.)", format="%d szt", alignment="center"),
                "Faktyczny_Obrot_Netto": str_web.column_config.NumberColumn("Faktyczny Obrót Netto", format="%.2f PLN", alignment="center"),
                "Faktyczna_Marza_Netto": str_web.column_config.NumberColumn("Faktyczna Marża Netto", format="%.2f PLN", alignment="center")
            }
        )

        # --- EKSPORT DO EXCELA (Z WIERSZEM TOTAL DOPASOWANYM DO NOWEJ KOLEJNOŚCI) ---
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            wynik_excel = wynik.rename(columns={'Dni_Zlogu': nazwa_kolumny_zlogu})
            wynik_excel.to_excel(writer, index=False, sheet_name="Wynik Weryfikacji")
            worksheet = writer.sheets["Wynik Weryfikacji"]

            wiersz_total_idx = worksheet.max_row + 1

            worksheet.cell(row=wiersz_total_idx, column=1, value="TOTAL")
            worksheet.cell(row=wiersz_total_idx, column=2, value=int(srednie_dni_zlogu) if ma_kolumne_zlogu else 0)
            worksheet.cell(row=wiersz_total_idx, column=5, value=int(suma_paczek_mm))
            worksheet.cell(row=wiersz_total_idx, column=6, value=int(suma_sprzedanych_sztuk))
            worksheet.cell(row=wiersz_total_idx, column=7, value=float(total_obrot))
            worksheet.cell(row=wiersz_total_idx, column=8, value=float(total_marza))

            font_total = Font(name="Segoe UI", size=10, bold=True)
            border_top_thin = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))

            for col_idx in range(1, 9):
                cell = worksheet.cell(row=wiersz_total_idx, column=col_idx)
                cell.font = font_total
                cell.border = border_top_thin
                if col_idx in [2, 5, 6, 7, 8]:
                    cell.alignment = Alignment(horizontal="center")
                if col_idx in [7, 8]:
                    cell.number_format = '#,##0.00" PLN"'
                elif col_idx in [5, 6]:
                    cell.number_format = '#,##0" szt"'
                elif col_idx == 2:
                    cell.number_format = '#,##0" dni"' if ma_kolumne_zlogu else '"-"'

            dopasuj_szerokosci_kolumn(worksheet, margines=3, min_szerokosc=14)

        str_web.download_button(
            label="📥 Pobierz wynik weryfikacji do Excela",
            data=output.getvalue(),
            file_name=f"Weryfikacja_MM_{data_start}_{data_koniec}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
