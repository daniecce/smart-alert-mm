import streamlit as str_web
import pandas as pd
import re
from views.check_sales import wyswietl_weryfikator_sprzedazy

# --- IMPORTY NOWYCH MODUŁÓW ---
from config.styles import zaaplikuj_stylizacje
from core.data_loader import zaladuj_i_parsuj_dane
from views.main_report import wyswietl_ekran_glowny, stylizuj_matryce_logistycznie
from views.size_mask import wyswietl_karte_rozmiarowa

str_web.set_page_config(page_title="Smart Alert MM", layout="wide")

str_web.markdown(
    """
    <style>
        /* Twarde podciągnięcie głównego kontenera raportu */
        .block-container {
            padding-top: 0rem !important;
            padding-bottom: 0rem !important;
        }
        
        /* Wyzerowanie wewnętrznych marginesów panelu bocznego (likwiduje pas nad przyciskiem) */
        [data-testid="stSidebarUserContent"] {
            padding-top: 0rem !important;
        }
        
        /* Likwidacja domyślnego pustego bloku Streamlit na samej górze paska bocznego */
        [data-testid="stSidebarNav"] {
            padding-top: 0rem !important;
            margin-top: 0rem !important;
        }
        
        /* Ukrycie fabrycznego, przezroczystego nagłówka aplikacji */
        header {
            visibility: hidden;
            height: 0px;
        }
        
        /* Dodatkowe ucięcie marginesu pierwszego elementu w sidebarze */
        .element-container:first-child {
            margin-top: 0rem !important;
        }
    </style>
    """,
    unsafe_allow_html=True
)

# Uruchomienie profesjonalnych stylów
zaaplikuj_stylizacje()
str_web.title("🔴 SYSTEM SMART ALERT MM")
str_web.markdown("**INTER MOTORS STOCK OPTIMIZER** | Wersja BI 1.2")

# --- FUNKCJE POMOCNICZE ARCHITEKTURY ---
def wyciagnij_rdzen_po_przecinku(pelna_nazwa):
    tekst = str(pelna_nazwa).strip()
    match = re.search(r'\b(rozmiar)\b', tekst, re.IGNORECASE)
    if match:
        rdzen = tekst[:match.start()].strip()
        return rdzen.rstrip(', ')
    if ',' in tekst:
        return tekst.split(',')[0].strip()
    return tekst

def wyciagnij_czysty_rozmiar(pelna_nazwa):
    tekst = str(pelna_nazwa).upper()
    m = re.search(r'ROZMIAR\s+([A-Z0-9]+)', tekst)
    if m: return m.group(1).strip()
    return 'INNY'

def convert_df(df):
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Rekomendacje MM")
        workbook = writer.book
        
        worksheet = writer.sheets["Rekomendacje MM"]
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        ostatnia_litera = get_column_letter(max_col)
        
        tabela_excel = Table(displayName="TabelaRekomendacji", ref=f"A1:{ostatnia_litera}{max_row}")
        tabela_excel.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        worksheet.add_table(tabela_excel)
        
        font_naglowek = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        font_dane = Font(name="Segoe UI", size=10, bold=False)
        font_indeks = Font(name="Segoe UI", size=10, bold=True)
        
        fill_naglowek = PatternFill(start_color="E31B23", end_color="E31B23", fill_type="solid")
        fill_paski = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
        fill_alarm = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
        
        align_naglowek_srodek = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_lewa = Alignment(horizontal="left", vertical="center", wrap_text=False)
        align_srodek = Alignment(horizontal="center", vertical="center", wrap_text=False)
        border_cienka = Border(left=Side(style='thin', color='E0E0E0'), right=Side(style='thin', color='E0E0E0'), top=Side(style='thin', color='E0E0E0'), bottom=Side(style='thin', color='E0E0E0'))

        worksheet.row_dimensions[1].height = 80
        for cell in worksheet[1]:
            cell.font = font_naglowek
            cell.fill = fill_naglowek
            cell.alignment = align_naglowek_srodek
            cell.border = border_cienka
            
        # 🟢 Dynamiczne mapowanie kolumn na podstawie ich nazw w nagłówku, aby uniknąć błędów przesunięcia (shift)
        naglowki = [str(cell.value) for cell in worksheet[1]]
        
        idx_laczny_zlog = None
        for i, h in enumerate(naglowki, start=1):
            if "Łączna ilość w sklepach" in h:
                idx_laczny_zlog = i
                break

        for row_idx in range(2, max_row + 1):
            worksheet.row_dimensions[row_idx].height = 20
            
            # Bezpieczne pobranie wartości złogu z dynamicznie wyliczonej pozycji
            laczny_zlog_val = 0
            if idx_laczny_zlog:
                try:
                    laczny_zlog_val = int(worksheet.cell(row=row_idx, column=idx_laczny_zlog).value or 0)
                except (ValueError, TypeError):
                    laczny_zlog_val = 0

            for col_idx in range(1, max_col + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = font_dane
                cell.border = border_cienka
                
                # Dynamiczne formatowanie wyrównania tekstu w zależności od zawartości
                if col_idx in [1, 2, 3] or "Skąd zabrać" in str(naglowki[col_idx-1]):
                    cell.alignment = align_lewa
                else:
                    cell.alignment = align_srodek
 
                if col_idx == 1: 
                    cell.font = font_indeks
                if row_idx % 2 == 0: 
                    cell.fill = fill_paski
                
                # Alarmowanie dużego złogu na właściwej, dynamicznej kolumnie
                if idx_laczny_zlog and col_idx == idx_laczny_zlog and laczny_zlog_val >= 5:
                    cell.fill = fill_alarm
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="900C3F")

        # Dynamiczne ustawianie szerokości kolumn w zależności od długości zawartości nagłówka i danych
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            # Dajemy margines bezpieczeństwa na szerokość
            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    return output.getvalue()

def eksport_karty_maski(df):
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_flat = df.reset_index().copy() 
        df_flat.columns = ['_'.join([str(c) for c in col if c]).strip() if isinstance(col, tuple) else str(col) for col in df_flat.columns]
        df_flat.to_excel(writer, index=False, sheet_name="Maska Rozmiarowa")
        
        workbook = writer.book
        worksheet = writer.sheets["Maska Rozmiarowa"]
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        tabela_excel = Table(displayName="TabelaSiatkiRozmiarowej", ref=f"A1:{get_column_letter(max_col)}{max_row}")
        tabela_excel.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        worksheet.add_table(tabela_excel)
        
        font_naglowek = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        font_dane = Font(name="Segoe UI", size=10, bold=False)
        font_indeks = Font(name="Segoe UI", size=10, bold=True)
        fill_naglowek = PatternFill(start_color="E31B23", end_color="E31B23", fill_type="solid")
        fill_paski = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
        align_naglowek_srodek = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_lewa = Alignment(horizontal="left", vertical="center", wrap_text=False)
        align_srodek = Alignment(horizontal="center", vertical="center", wrap_text=False)
        border_cienka = Border(left=Side(style='thin', color='E0E0E0'), right=Side(style='thin', color='E0E0E0'), top=Side(style='thin', color='E0E0E0'), bottom=Side(style='thin', color='E0E0E0'))

        worksheet.row_dimensions[1].height = 40
        for cell in worksheet[1]:
            cell.font = font_naglowek
            cell.fill = fill_naglowek
            cell.alignment = align_naglowek_srodek
            cell.border = border_cienka
            
        for row_idx in range(2, max_row + 1):
            worksheet.row_dimensions[row_idx].height = 20
            for col_idx in range(1, max_col + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = font_dane
                cell.border = border_cienka
                cell.alignment = align_lewa if col_idx in [2, 3] else align_srodek
                if col_idx == 2: cell.font = font_indeks
                if row_idx % 2 == 0: cell.fill = fill_paski

        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            if col_idx == 1: worksheet.column_dimensions[col_letter].width = 10   
            elif col_idx == 2: worksheet.column_dimensions[col_letter].width = 22 
            elif col_idx == 3: worksheet.column_dimensions[col_letter].width = 45 
            elif col_idx == 4: worksheet.column_dimensions[col_letter].width = 16 
            elif col_idx == 5: worksheet.column_dimensions[col_letter].width = 14 
            else: worksheet.column_dimensions[col_letter].width = 13                                             

    return output.getvalue()

# --- STANY APLIKACJI W PAMIĘCI RAM ---
if 'dane_zaladowane' not in str_web.session_state:
    str_web.session_state['dane_zaladowane'] = False
    str_web.session_state['df_stany'] = None
    str_web.session_state['df_sprzedaz_surowa'] = None
    str_web.session_state['df_minmax'] = None
if 'aktywny_ekran' not in str_web.session_state:
    str_web.session_state['aktywny_ekran'] = "glowny"
if 'wybrany_rdzen_opisu' not in str_web.session_state:
    str_web.session_state['wybrany_rdzen_opisu'] = None
if 'klikniety_indeks_startowy' not in str_web.session_state:
    str_web.session_state['klikniety_indeks_startowy'] = None
if 'klikniety_cel' not in str_web.session_state:
    str_web.session_state['klikniety_cel'] = None
if 'klikniete_zrodla' not in str_web.session_state:
    str_web.session_state['klikniete_zrodla'] = None

if 'zrobione_transakcje' not in str_web.session_state: str_web.session_state['zrobione_transakcje'] = []
if 'wykonane_mm_rozmiary' not in str_web.session_state: str_web.session_state['wykonane_mm_rozmiary'] = []
if 'rejestr_przesuniec_mm' not in str_web.session_state: str_web.session_state['rejestr_przesuniec_mm'] = []

# --- ZINTEGROWANY SYSTEM INICJALIZACJI I RETENCJI FILTRÓW ---
DOMYSLNE_FILTRY = {
    'f_zakres_dni': 30,
    'f_cena_min': 300,
    'f_marza_min': 0,
    'f_popyt_min': 2,
    'f_wybrani_producenci': [],
    'f_wybrane_statusy': [],
    'f_wybrane_filie_zrodlo': [],
    'f_wybrane_filie_cel': []
}

for klucz, wartosc_domyslna in DOMYSLNE_FILTRY.items():
    backup_klucz = f"{klucz}_backup"
    
    if klucz not in str_web.session_state:
        # 1. Kiedy suwak znika, odzyskujemy wartość z bezpiecznego schowka (o ile w nim jest)
        if backup_klucz in str_web.session_state:
            str_web.session_state[klucz] = str_web.session_state[backup_klucz]
        # 2. Tylko przy pierwszym wejściu do aplikacji ustawiamy wartość startową
        else:
            str_web.session_state[klucz] = wartosc_domyslna
    else:
        # 3. Kiedy użytkownik operuje suwakami, na bieżąco kopiujemy stan do schowka
        str_web.session_state[backup_klucz] = str_web.session_state[klucz]


# Autostart automatycznego ładowania zasobów z dysku
if not str_web.session_state['dane_zaladowane']:
    zaladuj_i_parsuj_dane()

# --- MANAGER NAWIGACJI EKRANÓW ---
if str_web.session_state['dane_zaladowane']:
    if str_web.session_state['aktywny_ekran'] == "glowny":
        wyswietl_ekran_glowny(wyciagnij_rdzen_po_przecinku, convert_df)
    elif str_web.session_state['aktywny_ekran'] == "detal_maski":
        wyswietl_karte_rozmiarowa(
            stylizuj_matryce_logistycznie, wyciagnij_czysty_rozmiar, 
            wyciagnij_rdzen_po_przecinku, eksport_karty_maski
        )
    elif str_web.session_state['aktywny_ekran'] == "weryfikator":
        wyswietl_weryfikator_sprzedazy()

# --- STOPKA INFORMACYJNA ---
str_web.markdown("---")
str_web.markdown(
    """
    <div style="text-align: center; color: var(--text-color); opacity: 0.5; font-size: 12px; font-family: 'Segoe UI', sans-serif; padding-top: 10px;">
        © 2026 | System zaprojektowany i stworzony przez: <b>Daniel Cecelon</b> | Inter Cars S.A.
    </div>
    """, 
    unsafe_allow_html=True
)