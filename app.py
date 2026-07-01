"""Entry point: Streamlit page config, screen routing, session_state initialization."""

import streamlit as str_web
import pandas as pd
import re
from views.check_sales import wyswietl_weryfikator_sprzedazy

# --- IMPORTY NOWYCH MODUŁÓW ---
from config.styles import zaaplikuj_stylizacje
from core.data_loader import zaladuj_i_parsuj_dane
from core.excel_export import (
    dopasuj_szerokosci_kolumn, stylizuj_naglowek_marki, stylizuj_wiersze_danych, dodaj_tabele_stylu_lekkiego
)
from views.main_report import wyswietl_ekran_glowny, stylizuj_matryce_logistycznie
from views.size_mask import wyswietl_karte_rozmiarowa

str_web.set_page_config(page_title="Smart Alert MM", layout="wide")

str_web.markdown(
    """
    <style>
        /* Twarde podciągnięcie głównego kontenera raportu */
        .block-container {
            padding-top: 0rem !important;
            padding-bottom: 0rem !important;
        }

        /* Wyzerowanie wewnętrznych marginesów panelu bocznego (likwiduje pas nad przyciskiem) */
        [data-testid="stSidebarUserContent"] {
            padding-top: 0rem !important;
        }

        /* Likwidacja domyślnego pustego bloku Streamlit na samej górze paska bocznego */
        [data-testid="stSidebarNav"] {
            padding-top: 0rem !important;
            margin-top: 0rem !important;
        }

        /* Ukrycie fabrycznego, przezroczystego nagłówka aplikacji */
        header {
            visibility: hidden;
            height: 0px;
        }

        /* Dodatkowe ucięcie marginesu pierwszego elementu w sidebarze */
        .element-container:first-child {
            margin-top: 0rem !important;
        }
    </style>
    """,
    unsafe_allow_html=True
)

# Uruchomienie profesjonalnych stylów
zaaplikuj_stylizacje()
str_web.title("🔴 SYSTEM SMART ALERT MM")
str_web.markdown("**INTER MOTORS STOCK OPTIMIZER** | Wersja BI 1.2")

# --- FUNKCJE POMOCNICZE ARCHITEKTURY ---
def wyciagnij_rdzen_po_przecinku(pelna_nazwa):
    tekst = str(pelna_nazwa).strip()
    match = re.search(r'\b(rozmiar)\b', tekst, re.IGNORECASE)
    if match:
        rdzen = tekst[:match.start()].strip()
        return rdzen.rstrip(', ')
    if ',' in tekst:
        return tekst.split(',')[0].strip()
    return tekst

def wyciagnij_czysty_rozmiar(pelna_nazwa):
    tekst = str(pelna_nazwa).upper()
    m = re.search(r'ROZMIAR\s+([A-Z0-9]+)', tekst)
    if m: return m.group(1).strip()
    return 'INNY'

def convert_df(df):
    from io import BytesIO
    from openpyxl.styles import Font, PatternFill

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Rekomendacje MM")
        worksheet = writer.sheets["Rekomendacje MM"]
        max_row = worksheet.max_row

        dodaj_tabele_stylu_lekkiego(worksheet, "TabelaRekomendacji")
        stylizuj_naglowek_marki(worksheet, wysokosc_wiersza=80)

        # Dynamiczne mapowanie kolumn na podstawie ich nazw w nagłówku, aby uniknąć błędów przesunięcia (shift)
        naglowki = [str(cell.value) for cell in worksheet[1]]

        idx_laczny_zlog = None
        for i, h in enumerate(naglowki, start=1):
            if "Łączna ilość w sklepach" in h:
                idx_laczny_zlog = i
                break

        kolumny_wyrownane_w_lewo = {1, 2, 3} | {i for i, h in enumerate(naglowki, start=1) if "Skąd zabrać" in h}
        stylizuj_wiersze_danych(worksheet, kolumny_wyrownane_w_lewo, kolumna_pogrubiona=1)

        # Alarmowanie dużego złogu na właściwej, dynamicznej kolumnie (reguła biznesowa tego
        # raportu, nie generyczny styl - dlatego dokładana tu, a nie w core/excel_export.py)
        if idx_laczny_zlog:
            fill_alarm = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
            font_alarm = Font(name="Segoe UI", size=10, bold=True, color="900C3F")
            for row_idx in range(2, max_row + 1):
                try:
                    laczny_zlog_val = int(worksheet.cell(row=row_idx, column=idx_laczny_zlog).value or 0)
                except (ValueError, TypeError):
                    laczny_zlog_val = 0
                if laczny_zlog_val >= 5:
                    cell = worksheet.cell(row=row_idx, column=idx_laczny_zlog)
                    cell.fill = fill_alarm
                    cell.font = font_alarm

        dopasuj_szerokosci_kolumn(worksheet, margines=3, min_szerokosc=12)

    return output.getvalue()

def eksport_karty_maski(df):
    from io import BytesIO
    from openpyxl.utils import get_column_letter

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_flat = df.reset_index().copy()
        df_flat.columns = ['_'.join([str(c) for c in col if c]).strip() if isinstance(col, tuple) else str(col) for col in df_flat.columns]
        df_flat.to_excel(writer, index=False, sheet_name="Maska Rozmiarowa")

        worksheet = writer.sheets["Maska Rozmiarowa"]
        max_col = worksheet.max_column

        dodaj_tabele_stylu_lekkiego(worksheet, "TabelaSiatkiRozmiarowej")
        stylizuj_naglowek_marki(worksheet, wysokosc_wiersza=40)
        stylizuj_wiersze_danych(worksheet, kolumny_wyrownane_w_lewo={2, 3}, kolumna_pogrubiona=2)

        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            if col_idx == 1: worksheet.column_dimensions[col_letter].width = 10
            elif col_idx == 2: worksheet.column_dimensions[col_letter].width = 22
            elif col_idx == 3: worksheet.column_dimensions[col_letter].width = 45
            elif col_idx == 4: worksheet.column_dimensions[col_letter].width = 16
            elif col_idx == 5: worksheet.column_dimensions[col_letter].width = 14
            else: worksheet.column_dimensions[col_letter].width = 13

    return output.getvalue()

# --- STANY APLIKACJI W PAMIĘCI RAM ---
if 'dane_zaladowane' not in str_web.session_state:
    str_web.session_state['dane_zaladowane'] = False
    str_web.session_state['df_stany'] = None
    str_web.session_state['df_sprzedaz_surowa'] = None
    str_web.session_state['df_minmax'] = None
if 'aktywny_ekran' not in str_web.session_state:
    str_web.session_state['aktywny_ekran'] = "glowny"
if 'wybrany_rdzen_opisu' not in str_web.session_state:
    str_web.session_state['wybrany_rdzen_opisu'] = None
if 'klikniety_indeks_startowy' not in str_web.session_state:
    str_web.session_state['klikniety_indeks_startowy'] = None
if 'klikniety_cel' not in str_web.session_state:
    str_web.session_state['klikniety_cel'] = None
if 'klikniete_zrodla' not in str_web.session_state:
    str_web.session_state['klikniete_zrodla'] = None

if 'zrobione_transakcje' not in str_web.session_state: str_web.session_state['zrobione_transakcje'] = []
if 'wykonane_mm_rozmiary' not in str_web.session_state: str_web.session_state['wykonane_mm_rozmiary'] = []
if 'rejestr_przesuniec_mm' not in str_web.session_state: str_web.session_state['rejestr_przesuniec_mm'] = []

# --- ZINTEGROWANY SYSTEM INICJALIZACJI I RETENCJI FILTRÓW ---
DOMYSLNE_FILTRY = {
    'f_zakres_dni': 30,
    'f_cena_min': 300,
    'f_marza_min': 0,
    'f_popyt_min': 2,
    'f_wybrani_producenci': [],
    'f_wybrane_statusy': [],
    'f_wybrane_filie_zrodlo': [],
    'f_wybrane_filie_cel': []
}

for klucz, wartosc_domyslna in DOMYSLNE_FILTRY.items():
    backup_klucz = f"{klucz}_backup"

    if klucz not in str_web.session_state:
        # 1. Kiedy suwak znika, odzyskujemy wartość z bezpiecznego schowka (o ile w nim jest)
        if backup_klucz in str_web.session_state:
            str_web.session_state[klucz] = str_web.session_state[backup_klucz]
        # 2. Tylko przy pierwszym wejściu do aplikacji ustawiamy wartość startową
        else:
            str_web.session_state[klucz] = wartosc_domyslna
    else:
        # 3. Kiedy użytkownik operuje suwakami, na bieżąco kopiujemy stan do schowka
        str_web.session_state[backup_klucz] = str_web.session_state[klucz]


# Autostart automatycznego ładowania zasobów z dysku
if not str_web.session_state['dane_zaladowane']:
    zaladuj_i_parsuj_dane()

# --- MANAGER NAWIGACJI EKRANÓW ---
if str_web.session_state['dane_zaladowane']:
    if str_web.session_state['aktywny_ekran'] == "glowny":
        wyswietl_ekran_glowny(wyciagnij_rdzen_po_przecinku, convert_df)
    elif str_web.session_state['aktywny_ekran'] == "detal_maski":
        wyswietl_karte_rozmiarowa(
            stylizuj_matryce_logistycznie, wyciagnij_czysty_rozmiar,
            wyciagnij_rdzen_po_przecinku, eksport_karty_maski
        )
    elif str_web.session_state['aktywny_ekran'] == "weryfikator":
        wyswietl_weryfikator_sprzedazy()

# --- STOPKA INFORMACYJNA ---
str_web.markdown("---")
str_web.markdown(
    """
    <div style="text-align: center; color: var(--text-color); opacity: 0.5; font-size: 12px; font-family: 'Segoe UI', sans-serif; padding-top: 10px;">
        © 2026 | System zaprojektowany i stworzony przez: <b>Daniel Cecelon</b> | Inter Cars S.A.
    </div>
    """,
    unsafe_allow_html=True
)
