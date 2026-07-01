"""
Test parzystości wizualnej: sprawdza, że convert_df i eksport_karty_maski dają
identyczny plik Excela (wartości + STYLE komórek) po przejściu na wspólne helpery
z core/excel_export.py, w porównaniu ze starą wersją (kod sprzed konsolidacji,
skopiowany tu 1:1 jako punkt odniesienia).

Dane są syntetyczne (nie firmowe), więc ten test nie wymaga pomijania - działa
zawsze, także na maszynie bez folderu Smart_MM/.

Porównanie nie odbywa się na surowych bajtach pliku .xlsx (to archiwum ZIP,
metadane kontenera mogą się różnić mimo identycznej zawartości), tylko na modelu
obiektowym arkusza wczytanym przez openpyxl: wartości komórek, czcionki, wypełnienia,
obramowania, wyrównania, wysokości wierszy, szerokości kolumn i obiekt Table.
"""

from io import BytesIO

import openpyxl
import pandas as pd
import pytest
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from core.excel_export import (
    dopasuj_szerokosci_kolumn, stylizuj_naglowek_marki, stylizuj_wiersze_danych, dodaj_tabele_stylu_lekkiego
)


# --- STARE WERSJE (kopia 1:1 kodu sprzed konsolidacji, punkt odniesienia) ---

def stary_convert_df(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Rekomendacje MM")
        worksheet = writer.sheets["Rekomendacje MM"]
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        ostatnia_litera = get_column_letter(max_col)

        tabela_excel = Table(displayName="TabelaRekomendacji", ref=f"A1:{ostatnia_litera}{max_row}")
        tabela_excel.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        worksheet.add_table(tabela_excel)

        font_naglowek = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        font_dane = Font(name="Segoe UI", size=10, bold=False)
        font_indeks = Font(name="Segoe UI", size=10, bold=True)

        fill_naglowek = PatternFill(start_color="E31B23", end_color="E31B23", fill_type="solid")
        fill_paski = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
        fill_alarm = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")

        align_naglowek_srodek = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_lewa = Alignment(horizontal="left", vertical="center", wrap_text=False)
        align_srodek = Alignment(horizontal="center", vertical="center", wrap_text=False)
        border_cienka = Border(left=Side(style='thin', color='E0E0E0'), right=Side(style='thin', color='E0E0E0'), top=Side(style='thin', color='E0E0E0'), bottom=Side(style='thin', color='E0E0E0'))

        worksheet.row_dimensions[1].height = 80
        for cell in worksheet[1]:
            cell.font = font_naglowek
            cell.fill = fill_naglowek
            cell.alignment = align_naglowek_srodek
            cell.border = border_cienka

        naglowki = [str(cell.value) for cell in worksheet[1]]

        idx_laczny_zlog = None
        for i, h in enumerate(naglowki, start=1):
            if "Łączna ilość w sklepach" in h:
                idx_laczny_zlog = i
                break

        for row_idx in range(2, max_row + 1):
            worksheet.row_dimensions[row_idx].height = 20

            laczny_zlog_val = 0
            if idx_laczny_zlog:
                try:
                    laczny_zlog_val = int(worksheet.cell(row=row_idx, column=idx_laczny_zlog).value or 0)
                except (ValueError, TypeError):
                    laczny_zlog_val = 0

            for col_idx in range(1, max_col + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = font_dane
                cell.border = border_cienka

                if col_idx in [1, 2, 3] or "Skąd zabrać" in str(naglowki[col_idx-1]):
                    cell.alignment = align_lewa
                else:
                    cell.alignment = align_srodek

                if col_idx == 1:
                    cell.font = font_indeks
                if row_idx % 2 == 0:
                    cell.fill = fill_paski

                if idx_laczny_zlog and col_idx == idx_laczny_zlog and laczny_zlog_val >= 5:
                    cell.fill = fill_alarm
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="900C3F")

        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    return output.getvalue()


def stary_eksport_karty_maski(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_flat = df.reset_index().copy()
        df_flat.columns = ['_'.join([str(c) for c in col if c]).strip() if isinstance(col, tuple) else str(col) for col in df_flat.columns]
        df_flat.to_excel(writer, index=False, sheet_name="Maska Rozmiarowa")

        worksheet = writer.sheets["Maska Rozmiarowa"]
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        tabela_excel = Table(displayName="TabelaSiatkiRozmiarowej", ref=f"A1:{get_column_letter(max_col)}{max_row}")
        tabela_excel.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        worksheet.add_table(tabela_excel)

        font_naglowek = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        font_dane = Font(name="Segoe UI", size=10, bold=False)
        font_indeks = Font(name="Segoe UI", size=10, bold=True)
        fill_naglowek = PatternFill(start_color="E31B23", end_color="E31B23", fill_type="solid")
        fill_paski = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
        align_naglowek_srodek = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_lewa = Alignment(horizontal="left", vertical="center", wrap_text=False)
        align_srodek = Alignment(horizontal="center", vertical="center", wrap_text=False)
        border_cienka = Border(left=Side(style='thin', color='E0E0E0'), right=Side(style='thin', color='E0E0E0'), top=Side(style='thin', color='E0E0E0'), bottom=Side(style='thin', color='E0E0E0'))

        worksheet.row_dimensions[1].height = 40
        for cell in worksheet[1]:
            cell.font = font_naglowek
            cell.fill = fill_naglowek
            cell.alignment = align_naglowek_srodek
            cell.border = border_cienka

        for row_idx in range(2, max_row + 1):
            worksheet.row_dimensions[row_idx].height = 20
            for col_idx in range(1, max_col + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = font_dane
                cell.border = border_cienka
                cell.alignment = align_lewa if col_idx in [2, 3] else align_srodek
                if col_idx == 2: cell.font = font_indeks
                if row_idx % 2 == 0: cell.fill = fill_paski

        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            if col_idx == 1: worksheet.column_dimensions[col_letter].width = 10
            elif col_idx == 2: worksheet.column_dimensions[col_letter].width = 22
            elif col_idx == 3: worksheet.column_dimensions[col_letter].width = 45
            elif col_idx == 4: worksheet.column_dimensions[col_letter].width = 16
            elif col_idx == 5: worksheet.column_dimensions[col_letter].width = 14
            else: worksheet.column_dimensions[col_letter].width = 13

    return output.getvalue()


# --- NOWE WERSJE (dokładnie to, co jest teraz w app.py) ---

def nowy_convert_df(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Rekomendacje MM")
        worksheet = writer.sheets["Rekomendacje MM"]
        max_row = worksheet.max_row

        dodaj_tabele_stylu_lekkiego(worksheet, "TabelaRekomendacji")
        stylizuj_naglowek_marki(worksheet, wysokosc_wiersza=80)

        naglowki = [str(cell.value) for cell in worksheet[1]]
        idx_laczny_zlog = None
        for i, h in enumerate(naglowki, start=1):
            if "Łączna ilość w sklepach" in h:
                idx_laczny_zlog = i
                break

        kolumny_wyrownane_w_lewo = {1, 2, 3} | {i for i, h in enumerate(naglowki, start=1) if "Skąd zabrać" in h}
        stylizuj_wiersze_danych(worksheet, kolumny_wyrownane_w_lewo, kolumna_pogrubiona=1)

        if idx_laczny_zlog:
            fill_alarm = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
            font_alarm = Font(name="Segoe UI", size=10, bold=True, color="900C3F")
            for row_idx in range(2, max_row + 1):
                try:
                    laczny_zlog_val = int(worksheet.cell(row=row_idx, column=idx_laczny_zlog).value or 0)
                except (ValueError, TypeError):
                    laczny_zlog_val = 0
                if laczny_zlog_val >= 5:
                    cell = worksheet.cell(row=row_idx, column=idx_laczny_zlog)
                    cell.fill = fill_alarm
                    cell.font = font_alarm

        dopasuj_szerokosci_kolumn(worksheet, margines=3, min_szerokosc=12)

    return output.getvalue()


def nowy_eksport_karty_maski(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_flat = df.reset_index().copy()
        df_flat.columns = ['_'.join([str(c) for c in col if c]).strip() if isinstance(col, tuple) else str(col) for col in df_flat.columns]
        df_flat.to_excel(writer, index=False, sheet_name="Maska Rozmiarowa")

        worksheet = writer.sheets["Maska Rozmiarowa"]
        max_col = worksheet.max_column

        dodaj_tabele_stylu_lekkiego(worksheet, "TabelaSiatkiRozmiarowej")
        stylizuj_naglowek_marki(worksheet, wysokosc_wiersza=40)
        stylizuj_wiersze_danych(worksheet, kolumny_wyrownane_w_lewo={2, 3}, kolumna_pogrubiona=2)

        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            if col_idx == 1: worksheet.column_dimensions[col_letter].width = 10
            elif col_idx == 2: worksheet.column_dimensions[col_letter].width = 22
            elif col_idx == 3: worksheet.column_dimensions[col_letter].width = 45
            elif col_idx == 4: worksheet.column_dimensions[col_letter].width = 16
            elif col_idx == 5: worksheet.column_dimensions[col_letter].width = 14
            else: worksheet.column_dimensions[col_letter].width = 13

    return output.getvalue()


# --- PRÓBKI DANYCH ---

def przykladowy_raport_rekomendacji():
    return pd.DataFrame({
        'Indeks': ['A1', 'A2', 'A3', 'A4', 'A5'],
        'Producent': ['MARKA1', 'MARKA2', 'MARKA1', 'MARKA3', 'MARKA2'],
        'Nazwa Produktu': ['Produkt Alfa', 'Produkt Beta', 'Produkt Gamma', 'Produkt Delta', 'Produkt Epsilon'],
        'Cena Detal Netto': [499.0, 899.0, 320.0, 1200.0, 650.0],
        'Marża Netto': [50.0, 120.0, 30.0, 200.0, 80.0],
        'Status': ['Aktywny'] * 5,
        'Sklep': ['S01', 'S02', 'S03', 'S04', 'S05'],
        'PJM (30 dni)': [3, 5, 2, 8, 1],
        'MAX': [2, 4, 1, 6, 1],
        'Skąd zabrać (Sklepy bez sprzedaży w 30 dni)': ['S10', 'S11, S12', 'S13', 'S14', 'S15'],
        # celowo mix: >=5 (alarm) i <5 (bez alarmu), na parzystych i nieparzystych wierszach arkusza
        'Łączna ilość w sklepach bez sprzedaży': [10, 3, 7, 2, 6],
    })


def przykladowa_maska_rozmiarowa():
    kolumny = pd.MultiIndex.from_tuples([
        ('Info', 'Indeks'), ('Info', 'Nazwa'),
        ('CZM', 'Stan'), ('CZM', 'Sprzedaż'),
        ('S01', 'Stan'), ('S01', 'Sprzedaż'),
    ])
    dane = [
        ['A1', 'Produkt Alfa rozmiar S', 5, 0, 0, 3],
        ['A2', 'Produkt Alfa rozmiar M', 2, 0, 0, 1],
        ['A3', 'Produkt Alfa rozmiar L', 0, 0, 4, 0],
    ]
    return pd.DataFrame(dane, columns=kolumny)


# --- PORÓWNANIE MODELU OBIEKTOWEGO ARKUSZA ---

def _styl_komorki(cell):
    kolor_fontu = cell.font.color.rgb if cell.font and cell.font.color else None
    kolor_wypelnienia = cell.fill.fgColor.rgb if cell.fill and cell.fill.patternType else None
    border = cell.border

    def bok(side):
        if side is None or side.style is None:
            return None
        return (side.style, side.color.rgb if side.color else None)

    return (
        cell.value,
        cell.font.name if cell.font else None,
        cell.font.size if cell.font else None,
        cell.font.bold if cell.font else None,
        kolor_fontu,
        cell.fill.patternType if cell.fill else None,
        kolor_wypelnienia,
        bok(border.left) if border else None,
        bok(border.right) if border else None,
        bok(border.top) if border else None,
        bok(border.bottom) if border else None,
        cell.alignment.horizontal if cell.alignment else None,
        cell.alignment.vertical if cell.alignment else None,
        cell.alignment.wrap_text if cell.alignment else None,
    )


def zrzutuj_arkusz(worksheet):
    komorki = {
        (r, c): _styl_komorki(worksheet.cell(row=r, column=c))
        for r in range(1, worksheet.max_row + 1)
        for c in range(1, worksheet.max_column + 1)
    }
    wysokosci = {r: worksheet.row_dimensions[r].height for r in range(1, worksheet.max_row + 1)}
    szerokosci = {
        get_column_letter(c): worksheet.column_dimensions[get_column_letter(c)].width
        for c in range(1, worksheet.max_column + 1)
    }
    tabele = {
        nazwa: (
            worksheet.tables[nazwa].ref,
            worksheet.tables[nazwa].tableStyleInfo.name,
            worksheet.tables[nazwa].tableStyleInfo.showRowStripes,
        )
        for nazwa in worksheet.tables.keys()
    }
    return komorki, wysokosci, szerokosci, tabele


def porownaj_pliki_excel(bajty_stare, bajty_nowe, nazwa_arkusza):
    wb_stary = openpyxl.load_workbook(BytesIO(bajty_stare))
    wb_nowy = openpyxl.load_workbook(BytesIO(bajty_nowe))

    ws_stary = wb_stary[nazwa_arkusza]
    ws_nowy = wb_nowy[nazwa_arkusza]

    komorki_s, wysokosci_s, szerokosci_s, tabele_s = zrzutuj_arkusz(ws_stary)
    komorki_n, wysokosci_n, szerokosci_n, tabele_n = zrzutuj_arkusz(ws_nowy)

    assert komorki_s == komorki_n, "Różnice w wartościach/stylach komórek"
    assert wysokosci_s == wysokosci_n, "Różnice w wysokościach wierszy"
    assert szerokosci_s == szerokosci_n, "Różnice w szerokościach kolumn"
    assert tabele_s == tabele_n, "Różnice w obiekcie Table (nazwa/zakres/styl)"


def test_convert_df_identyczny_po_konsolidacji():
    """convert_df po przejściu na core.excel_export musi dać identyczny plik jak wersja sprzed konsolidacji."""
    df = przykladowy_raport_rekomendacji()
    bajty_stare = stary_convert_df(df)
    bajty_nowe = nowy_convert_df(df)
    porownaj_pliki_excel(bajty_stare, bajty_nowe, "Rekomendacje MM")


def test_eksport_karty_maski_identyczny_po_konsolidacji():
    """eksport_karty_maski po przejściu na core.excel_export musi dać identyczny plik jak wersja sprzed konsolidacji."""
    df = przykladowa_maska_rozmiarowa()
    bajty_stare = stary_eksport_karty_maski(df)
    bajty_nowe = nowy_eksport_karty_maski(df)
    porownaj_pliki_excel(bajty_stare, bajty_nowe, "Maska Rozmiarowa")
